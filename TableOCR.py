import cv2
import numpy as np
import os
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter
from google.cloud import vision

# ------------------ Google Credential Setup ------------------

def setup_google_vision_client():
    service_account_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not service_account_json:
        raise EnvironmentError("❌ GOOGLE_SERVICE_ACCOUNT_JSON not set!")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as temp_file:
        temp_file.write(service_account_json.encode())
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = temp_file.name

    return vision.ImageAnnotatorClient()

# ------------------ Main Function: Table Extraction and Excel ------------------

def extract_table_and_save_excel(image_path):
    try:
        client = setup_google_vision_client()

        image = cv2.imread(image_path)
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        binary = cv2.adaptiveThreshold(~gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 15, -2)

        # Morphological table grid detection
        scale = 20
        horizontal = binary.copy()
        vertical = binary.copy()

        horizontalsize = horizontal.shape[1] // scale
        horizontal_structure = cv2.getStructuringElement(cv2.MORPH_RECT, (horizontalsize, 1))
        horizontal = cv2.erode(horizontal, horizontal_structure)
        horizontal = cv2.dilate(horizontal, horizontal_structure)

        verticalsize = vertical.shape[0] // scale
        vertical_structure = cv2.getStructuringElement(cv2.MORPH_RECT, (1, verticalsize))
        vertical = cv2.erode(vertical, vertical_structure)
        vertical = cv2.dilate(vertical, vertical_structure)

        mask = cv2.add(horizontal, vertical)
        mask = cv2.dilate(mask, np.ones((3, 3), np.uint8))
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        boxes = [cv2.boundingRect(cnt) for cnt in contours if cv2.boundingRect(cnt)[2] > 50 and cv2.boundingRect(cnt)[3] > 20]
        boxes = sorted(boxes, key=lambda b: (b[1], b[0]))

        # Group boxes into rows
        rows = []
        row_tolerance = 10
        for box in boxes:
            x, y, w, h = box
            added = False
            for row in rows:
                if abs(row[0][1] - y) < row_tolerance:
                    row.append(box)
                    added = True
                    break
            if not added:
                rows.append([box])

        for row in rows:
            row.sort(key=lambda b: b[0])  # sort within row by x

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Table"

        for row_idx, row in enumerate(rows, start=1):
            for col_idx, (x, y, w, h) in enumerate(row, start=1):
                cell_img = image[y:y+h, x:x+w]
                _, encoded_image = cv2.imencode('.jpg', cell_img)
                content = encoded_image.tobytes()
                vision_image = vision.Image(content=content)
                response = client.text_detection(image=vision_image)
                texts = response.text_annotations
                text = texts[0].description.strip().replace('\n', ' ') if texts else ""
                ws.cell(row=row_idx, column=col_idx).value = text

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4

        # Save Excel file
        output_dir = os.path.join(os.path.dirname(image_path), 'outputexcel')
        os.makedirs(output_dir, exist_ok=True)
        output_excel_path = os.path.join(output_dir, os.path.splitext(os.path.basename(image_path))[0] + '_table.xlsx')
        wb.save(output_excel_path)

        print(f"✅ Excel file saved at: {output_excel_path}")
        return output_excel_path

    except Exception as e:
        print(f"❌ Failed: {e}")
        return None

# ------------------ Alias for compatibility ------------------

def extract_text_and_generate_csv(image_path):
    """Alias wrapper for legacy usage."""
    return extract_table_and_save_excel(image_path)

# ------------------ CLI Usage ------------------

if __name__ == "__main__":
    test_image = "your_image.png"  # Replace with your image path
    extract_text_and_generate_csv(test_image)
